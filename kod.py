import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime

primary_color = "#00AADB"

st.set_page_config(
    page_title="Total Reach 360°",
    page_icon=":bar_chart:",
    layout="wide",
)

pd.set_option('display.float_format', '{:.0f}'.format)


df = pd.read_excel('TBR360_g.xlsx')
tematyka = pd.read_excel('kat.xlsx')

tematyka_lista = tematyka['kat'].unique()
wskaźniki_lista = ['Druk i E-wydania', 'www PC', 'www Mobile', 'www', 'Total Reach 360°']
miesiące_lista = [1, 2, 3, 4, 5, 6, 7, 8, 9]

strony = pd.read_excel('strony.xlsx')
tematyka_legenda_dict = dict(zip(strony['Pismo'], strony['Strona']))
wydawca_legenda_dict = dict(zip(tematyka['tytuł'], tematyka['wydawca']))

st.markdown("<h1 style='margin-top: -80px; text-align: center;'>Total Reach 360°</h1>", unsafe_allow_html=True)

selected_miesiace = [341,342,343,344,345,346,347,348,349,350,351,352]

Płeć = st.radio("Wybierz płeć:", ['Wszyscy', 'Kobiety', 'Mężczyźni'], horizontal=True, index =0)

Wiek = st.multiselect("Wybierz grupę wiekową:", ['15-24', '25-34', '35-44', '45-59', '60-75'], default=['15-24', '25-34', '35-44', '45-59', '60-75'])

Grupa = st.radio("Wybierz grupę celową:", ['Wszyscy', 'Dochód gospodarstwa ponad 5 tys.', 'Dochód ponad 2 tys.',
                                           'Mieszkańcy miast powyżej 50 tys.', 'Osoby z dziećmi w wieku 0-14'], horizontal=True, index =0)

col1, col2 = st.columns([2.2,1])
with col1:
    selected_tematyki = st.multiselect("Określ grupy pism:", tematyka_lista, default=tematyka_lista)
with col2:
    wyszukiwarka = st.text_input("Wyszukaj markę prasową:",  "", key="placeholder")
if selected_tematyki == []:
    selected_tematyki = tematyka_lista

if Płeć == 'Wszyscy' and set(Wiek)== {'15-24', '25-34', '35-44', '45-59', '60-75'} and Grupa == 'Wszyscy': 
    try:
        if z==1:
            estymacja = st.radio("Określ sposób prezentowania danych:", ['Estymacja na populację', 'Zasięg (%)'], horizontal=True, index = 1)
    except:
        estymacja = st.radio("Określ sposób prezentowania danych:", ['Estymacja na populację', 'Zasięg (%)'], horizontal=True, index = 0)
else:
    try:
        if z==1:
            estymacja = st.radio("Określ sposób prezentowania danych:", ['Estymacja na populację', 'Zasięg (%)', 'Affinity index'], horizontal=True, index = 1)
    except:
        estymacja = st.radio("Określ sposób prezentowania danych:", ['Estymacja na populację', 'Zasięg (%)', 'Affinity index'], horizontal=True, index = 0)

if estymacja == 'Zasięg (%)':
    st.session_state["z"] = 1
else:
    st.session_state["z"] = 0

if estymacja == 'Affinity index' and Płeć == 'Wszyscy' and set(Wiek)== {'15-24', '25-34', '35-44', '45-59', '60-75'} and Grupa == 'Wszyscy':
    estymacja = 'Estymacja na populację'

www_option = st.radio("Określ zakres danych www:", ['Total Reach 360° (Druk i E-Wydania, www PC oraz www Mobile)', 'Total Reach 360° (Druk i E-Wydania, www)',
                                                    'Druk i E-wydania', 'www', 'www PC', 'www Mobile'], horizontal=True, index =0)

col1, col2= st.columns([1,3])

with col1:
    show_wydawca = st.checkbox("Pokaż wydawców", value=False)

if www_option == 'Total Reach 360° (Druk i E-Wydania, www PC oraz www Mobile)' or www_option == 'Total Reach 360° (Druk i E-Wydania, www)': 
    with col2:
        show_wspolczytelnictwo = st.checkbox("Pokaż współczytelnictwo", value=False)
else:
    show_wspolczytelnictwo = False


wyniki = pd.DataFrame()
wyniki_cal = pd.DataFrame()

for i in selected_tematyki:
    pisma_lista = tematyka[tematyka['kat'] == i]['tytuł'].to_list()
    for j in pisma_lista:
        for k in wskaźniki_lista:
            if k != 'Total Reach 360°':
                df_g = df.copy()
                if Płeć == 'Kobiety':
                    df_g = df_g[df_g['P']=='K']
                if Płeć == 'Mężczyźni':
                    df_g = df_g[df_g['P']=='M']
                wartosci_numeryczne = {'15-24': 1, '25-34': 2, '35-44': 3, '45-59': 4, '60-75': 5}
                Wiek_num = [wartosci_numeryczne[grupa] for grupa in Wiek]
                df_g = df_g[df_g['W'].isin(Wiek_num)]
               
                wyniki.loc[j, k] = df_g[(df_g['tytuł'] == j) & (df_g['WSKAŹNIK'] == k) & (df_g['WAVE'].between(selected_miesiace[0], selected_miesiace[-1]))]['WYNIK'].sum()
                if Grupa == 'Dochód gospodarstwa ponad 5 tys.':
                    if k == 'Druk i E-wydania':
                        wyniki.loc[j, k] = wyniki.loc[j, k] * float(df[(df['tytuł'] == j) & (df['WSKAŹNIK'] == 'w1_druk')]['WYNIK'])
                    else:
                        wyniki.loc[j, k] = wyniki.loc[j, k] * float(df[(df['tytuł'] == j) & (df['WSKAŹNIK'] == 'w1_www')]['WYNIK']) 
                if Grupa == 'Dochód ponad 2 tys.':
                    if k == 'Druk i E-wydania':
                        wyniki.loc[j, k] = wyniki.loc[j, k] * float(df[(df['tytuł'] == j) & (df['WSKAŹNIK'] == 'w2_druk')]['WYNIK']) 
                    else:
                        wyniki.loc[j, k] = wyniki.loc[j, k] * float(df[(df['tytuł'] == j) & (df['WSKAŹNIK'] == 'w2_www')]['WYNIK']) 
                if Grupa == 'Mieszkańcy miast powyżej 50 tys.':
                    if k == 'Druk i E-wydania':
                        wyniki.loc[j, k] = wyniki.loc[j, k] * float(df[(df['tytuł'] == j) & (df['WSKAŹNIK'] == 'w3_druk')]['WYNIK']) 
                    else:
                        wyniki.loc[j, k] = wyniki.loc[j, k] * float(df[(df['tytuł'] == j) & (df['WSKAŹNIK'] == 'w3_www')]['WYNIK']) 
                if Grupa == 'Osoby z dziećmi w wieku 0-14':
                    if k == 'Druk i E-wydania':
                        wyniki.loc[j, k] = wyniki.loc[j, k] * float(df[(df['tytuł'] == j) & (df['WSKAŹNIK'] == 'w4_druk')]['WYNIK']) 
                    else:
                        wyniki.loc[j, k] = wyniki.loc[j, k] * float(df[(df['tytuł'] == j) & (df['WSKAŹNIK'] == 'w4_www')]['WYNIK']) 

                                            
                wyniki_cal.loc[j, k] = df[(df['tytuł'] == j) & (df['WSKAŹNIK'] == k) & (df['WAVE'].between(selected_miesiace[0], selected_miesiace[-1]))]['WYNIK'].sum()
            else:
                wyniki.loc[j, k] = max(wyniki.loc[j, 'Druk i E-wydania'], (1 - float(df[(df['tytuł'] == j) & (df['WSKAŹNIK'] == 'współczytelnictwo')]['WYNIK'])) * wyniki.loc[j, 'Druk i E-wydania'] + wyniki.loc[j, 'www'])
                wyniki_cal.loc[j, k] = max(wyniki_cal.loc[j, 'Druk i E-wydania'], (1 - float(df[(df['tytuł'] == j) & (df['WSKAŹNIK'] == 'współczytelnictwo')]['WYNIK'])) * wyniki_cal.loc[j, 'Druk i E-wydania'] + wyniki_cal.loc[j, 'www'])


wyniki = wyniki[wyniki.index.str.contains(wyszukiwarka, case=False, na=False)]
wyniki_cal = wyniki_cal[wyniki_cal.index.str.contains(wyszukiwarka, case=False, na=False)]

if estymacja == 'Affinity index':
    wyniki = wyniki / wyniki_cal  * 29545225
    wyniki = wyniki.fillna(0)

if www_option == 'Druk i E-wydania':
    wyniki = wyniki.sort_values('Druk i E-wydania', ascending=False)
elif www_option == 'www':
    wyniki = wyniki.sort_values('www', ascending=False)
elif www_option == 'www PC':
    wyniki = wyniki.sort_values('www PC', ascending=False)
elif www_option == 'www Mobile':
    wyniki = wyniki.sort_values('www Mobile', ascending=False)
else:
    wyniki = wyniki.sort_values('Total Reach 360°', ascending=False)


suma = 0 
if (Płeć == 'Kobiety' or Płeć == 'Wszyscy') and 1 in Wiek_num:
    suma += 1812738
if (Płeć == 'Mężczyźni' or Płeć == 'Wszyscy') and 1 in Wiek_num:
    suma += 1893705
if (Płeć == 'Kobiety' or Płeć == 'Wszyscy') and 2 in Wiek_num:
    suma += 2387225
if (Płeć == 'Mężczyźni' or Płeć == 'Wszyscy') and 2 in Wiek_num:
    suma += 2440240
if (Płeć == 'Kobiety' or Płeć == 'Wszyscy') and 3 in Wiek_num:
    suma += 3054633
if (Płeć == 'Mężczyźni' or Płeć == 'Wszyscy') and 3 in Wiek_num:
    suma += 3054845
if (Płeć == 'Kobiety' or Płeć == 'Wszyscy') and 4 in Wiek_num:
    suma += 3762817
if (Płeć == 'Mężczyźni' or Płeć == 'Wszyscy') and 4 in Wiek_num:
    suma += 3596251
if (Płeć == 'Kobiety' or Płeć == 'Wszyscy') and 5 in Wiek_num:
    suma += 4205663
if (Płeć == 'Mężczyźni' or Płeć == 'Wszyscy') and 5 in Wiek_num:
    suma += 3337109

if Grupa == 'Dochód gospodarstwa ponad 5 tys.':
    suma = suma * 0.42544612
if Grupa == 'Dochód ponad 2 tys.':
    suma = suma * 0.750720851
if Grupa == 'Mieszkańcy miast powyżej 50 tys.':
    suma = suma * 0.346943213
if Grupa == 'Osoby z dziećmi w wieku 0-14':
    suma = suma * 0.230303669

suma = int(round(suma,0))

if estymacja == 'Zasięg (%)' or  estymacja == 'Affinity index' :
    wyniki = wyniki / suma * 100
    wyniki = wyniki.round(2)

if show_wspolczytelnictwo:
    wyniki['Współczytelnictwo'] = wyniki['Druk i E-wydania'] + wyniki['www'] - wyniki['Total Reach 360°']

wyniki.replace(0, np.nan, inplace=True)

wyniki_sformatowane = wyniki.applymap(lambda x: '{:,.2f}%'.format(x).replace('.', ',') if not pd.isna(x) and estymacja == 'Zasięg (%)' else '{:,.0f}'.format(x).replace(',', ' ') if not pd.isna(x) else x)

wyniki_sformatowane = wyniki_sformatowane.fillna('-')

if www_option ==  'Total Reach 360° (Druk i E-Wydania, www)':
    del wyniki_sformatowane['www PC']
    del wyniki_sformatowane['www Mobile']
elif www_option == 'Total Reach 360° (Druk i E-Wydania, www PC oraz www Mobile)':
    del wyniki_sformatowane['www']
else:
    del wyniki_sformatowane['Total Reach 360°']
    if www_option ==  'Druk i E-wydania':
        del wyniki_sformatowane['www PC']
        del wyniki_sformatowane['www Mobile']
        del wyniki_sformatowane['www']
    elif www_option ==  'www':
        del wyniki_sformatowane['www PC']
        del wyniki_sformatowane['www Mobile']
        del wyniki_sformatowane['Druk i E-wydania']
    elif www_option ==  'www PC':
        del wyniki_sformatowane['www']
        del wyniki_sformatowane['www Mobile']
        del wyniki_sformatowane['Druk i E-wydania']
    elif www_option ==  'www Mobile':
        del wyniki_sformatowane['www PC']
        del wyniki_sformatowane['www']
        del wyniki_sformatowane['Druk i E-wydania']

if www_option ==  'Total Reach 360° (Druk i E-Wydania, www)':
    wyniki_sformatowane = wyniki_sformatowane.astype('object').fillna('-')
elif www_option ==  'www':
    wyniki_sformatowane.dropna(subset = ['www'], inplace=True)


wyniki_sformatowane = wyniki_sformatowane.reset_index()
wyniki_sformatowane.columns = ['Marka prasowa'] + list(wyniki_sformatowane.columns[1:])
wyniki_sformatowane['Wydawca'] = wyniki_sformatowane['Marka prasowa'].map(wydawca_legenda_dict)
new_column_order = ['Marka prasowa', 'Wydawca'] + list(wyniki_sformatowane.columns[1:-1])
#wyniki_sformatowane['Marka prasowa'] = wyniki_sformatowane['Marka prasowa'].apply(lambda x: f"[{x}](https://www.pbc.pl/badany-tytul/{x.lower().replace(' ', '-')}/)")


wyniki_sformatowane = wyniki_sformatowane[new_column_order]
wyniki_sformatowane.index+=1


if show_wydawca == False:
    del wyniki_sformatowane['Wydawca']

wyniki_sformatowane_2 = wyniki_sformatowane.copy()

if 'Total Reach 360°' in wyniki_sformatowane.columns :
    wyniki_sformatowane_styled = wyniki_sformatowane.style.set_table_styles([
    {'selector': 'table', 'props': [('text-align', 'center')]},
    {'selector': 'th', 'props': [('text-align', 'center')]},
    {'selector': 'td', 'props': [('text-align', 'center')]},
    {'selector': 'th.col0, td.col0', 'props': [('text-align', 'left')]}  # Wyrównaj pierwszą kolumnę do lewej
]).set_properties(
    subset=['Total Reach 360°'], 
    **{'font-weight': 'bold'})

else:
    wyniki_sformatowane_styled = wyniki_sformatowane.style.set_table_styles([
    {'selector': 'table', 'props': [('text-align', 'center')]},
    {'selector': 'th', 'props': [('text-align', 'center')]},
    {'selector': 'td', 'props': [('text-align', 'center')]},
    {'selector': 'th.col0, td.col0', 'props': [('text-align', 'left')]}  # Wyrównaj pierwszą kolumnę do lewej
])

def make_clickable(tytul):
    link = f"https://www.pbc.pl/badany-tytul/{tytul.lower().replace(' ', '-').replace('ó', 'o').replace('ś', 's').replace('ć', 'c').replace('ł', 'l').replace('ł', 'l').replace('ń', 'n').replace('ó', 'o').replace('ż', 'z')}/"
    return f'<a target="_blank" href="{link}">{tytul}</a>'

wyniki_sformatowane['Marka prasowa'] = wyniki_sformatowane['Marka prasowa'].apply(make_clickable)

html_table = wyniki_sformatowane_styled.to_html()


html_table = f"<div style='margin: auto;'>{html_table}</div>"

styled_table = f"""
<style>
    table {{
        width: 100%;
        margin: auto;
        overflow-x: auto;
    }}
    th, td {{
        padding: 10px;
        text-align: left;
        border: 1px solid #ddd;
        white-space: nowrap;  /* Unikaj przerywania tekstu na wielu linijkach */
    }}
</style>
{html_table}
"""

# Wyświetl sformatowaną tabelę
st.markdown(styled_table, unsafe_allow_html=True)

tekst = 'Badane marki:'
for pismo in wyniki.index.unique():
    try:
        tekst = f'{tekst} {pismo} i {tematyka_legenda_dict[pismo]},'
    except:
        pass

st.markdown(f"""<div style="font-size:12px">Statystyki: Zasięg CCS i Estymacja na populację, Populacja w wybranej grupie celowej =  {suma}</div>""", unsafe_allow_html=True)

st.markdown("""<div style="font-size:12px">Fale: 1-12/2023</div>""", unsafe_allow_html=True)


st.markdown("""<div style="font-size:12px">Dane CCS: Druk, E-wydania, Współczytelnictwo – Badanie PBC „Zanagażowanie w reklamę” ,
www, www PC, www mobile – Realusers (RU) PBI/Gemius</div>""", unsafe_allow_html=True)
            
st.markdown(f"""<div style="font-size:12px">{tekst}</div>""", unsafe_allow_html=True)

st.markdown("""<div style="font-size:12px">Definicje: www.pbc.pl/wskazniki/</div>""", unsafe_allow_html=True)

plik_wejsciowy = "szablon.xlsx"
arkusz = openpyxl.load_workbook(plik_wejsciowy)
arkusz.active['A3'] = f'Data wykonania raportu: {datetime.now().strftime("%d.%m.%Y")}'
arkusz.active['A85'] = tekst

wyniki_sformatowane_2 = wyniki_sformatowane_2.applymap(lambda x: int(x) if isinstance(x, (int, float)) else x)

for i, nazwa_kolumny in enumerate(wyniki_sformatowane_2.columns, start=1):
    arkusz.active.cell(row=5, column=i, value=nazwa_kolumny)

for i, wiersz in enumerate(wyniki_sformatowane_2.iloc[0:].itertuples(), start=6):
    for j, wartosc in enumerate(wiersz[1:], start=1):
        arkusz.active.cell(row=i, column=j, value=wartosc)

for col_index in range(8, wyniki_sformatowane_2.shape[1], -1):
    arkusz.active.delete_cols(col_index)

for row_index in range(80, wyniki_sformatowane_2.shape[0] + 5, -1):
    arkusz.active.delete_rows(row_index)

for wiersz in range(6, 81):
    for kolumna in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        komorka = arkusz.active[f'{kolumna}{wiersz}']
        try:
            komorka.value = int(komorka.value.replace(' ', ''))
        except:
            pass

plik_wyjsciowy = f"TR_{datetime.now().strftime('%d.%m.%Y')}.xlsx"
arkusz.save(plik_wyjsciowy)

st.download_button(
    label="Zapisz raport do pliku",
    data=open(plik_wyjsciowy, 'rb').read(),
    file_name=plik_wyjsciowy,
    mime="application/vnd.ms-excel"
)